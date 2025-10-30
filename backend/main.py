from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import io
import os
from datetime import datetime
import tempfile
import logging
from openpyxl import load_workbook
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Bespoke Financial Model API")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Path to the template model file
TEMPLATE_MODEL_PATH = os.path.join(os.path.dirname(__file__), "Bespoke Model - US - v2.xlsm")

@app.on_event("startup")
async def startup_event():
    """Check template file on startup"""
    logger.info(f"Backend starting up...")
    logger.info(f"Looking for template at: {TEMPLATE_MODEL_PATH}")
    logger.info(f"Template exists: {os.path.exists(TEMPLATE_MODEL_PATH)}")
    if os.path.exists(TEMPLATE_MODEL_PATH):
        logger.info(f"Template file size: {os.path.getsize(TEMPLATE_MODEL_PATH)} bytes")
    else:
        logger.warning(f"⚠️ Template file not found! Please add 'Bespoke Model - US - v2.xlsm' to the backend directory")

def process_bespoke_model(input_file_bytes: bytes) -> bytes:
    """
    Process the Bespoke Input Sheet and generate the output model
    """
    try:
        logger.info(f"[v0] Starting to process input file ({len(input_file_bytes)} bytes)")
        
        # Check if template exists
        if not os.path.exists(TEMPLATE_MODEL_PATH):
            raise FileNotFoundError(
                f"Template file 'Bespoke Model - US - v2.xlsm' not found at {TEMPLATE_MODEL_PATH}. "
                "Please add the template file to the backend directory."
            )
        
        logger.info(f"[v0] Loading input workbook...")
        # Load the input workbook
        input_wb = load_workbook(io.BytesIO(input_file_bytes), keep_vba=True, data_only=True)
        
        # Check if the required sheet exists
        if "Sales Team Input Sheet" not in input_wb.sheetnames:
            logger.error(f"[v0] Available sheets: {input_wb.sheetnames}")
            raise ValueError("Input file must contain a 'Sales Team Input Sheet' worksheet")
        
        ws_input = input_wb["Sales Team Input Sheet"]
        logger.info(f"[v0] Reading values from input sheet...")
        
        # Read values from the input sheet
        address_raw = str(ws_input["F7"].value or "").strip()
        additional_info = str(ws_input["F9"].value or "").strip()
        latitude = ws_input["F13"].value
        longitude = ws_input["F15"].value
        sqft = ws_input["F29"].value
        brand = ws_input["F23"].value
        market_rent = ws_input["F37"].value
        yes_no_value = ws_input["F54"].value
        num_floors = ws_input["F56"].value
        
        logger.info(f"[v0] Read address: {address_raw}")
        logger.info(f"[v0] Read sqft: {sqft}, market_rent: {market_rent}, floors: {num_floors}")
        
        if not address_raw:
            raise ValueError("Address in cell F7 is empty. Please fill it in.")
        
        # Replace "Dr" → "Drive" and "Blvd" → "Boulevard" for model workbook
        address_clean = address_raw.replace("Dr", "Drive").replace("Blvd", "Boulevard")
        
        logger.info(f"[v0] Loading template model from {TEMPLATE_MODEL_PATH}...")
        # Load the template model workbook
        model_wb = load_workbook(TEMPLATE_MODEL_PATH, keep_vba=True)
        
        # Check if the required sheet exists in the model
        if "Sales Team Input Sheet" not in model_wb.sheetnames:
            logger.error(f"[v0] Available sheets in template: {model_wb.sheetnames}")
            raise ValueError("Template file must contain a 'Sales Team Input Sheet' worksheet")
        
        ws_model = model_wb["Sales Team Input Sheet"]
        logger.info(f"[v0] Updating model with input data...")
        
        # Update building address
        ws_model["E6"] = address_clean
        
        # Update coordinates
        if latitude is not None:
            ws_model["E12"] = latitude
        if longitude is not None:
            ws_model["E14"] = longitude
        
        # Update square footage (F29 → E34)
        if sqft is not None:
            ws_model["E34"] = sqft
        
        # Handle F37 (market rent) → dropdown selection in K10
        if market_rent is not None:
            try:
                market_rent_float = float(market_rent)
                if market_rent_float == 15:
                    ws_model["K10"] = "15 - 20"
                elif market_rent_float == 20:
                    ws_model["K10"] = "20 - 25"
                else:
                    ws_model["K10"] = market_rent
            except (ValueError, TypeError):
                ws_model["K10"] = market_rent
        
        # Copy F54 text into K34
        if yes_no_value is not None:
            ws_model["K34"] = str(yes_no_value).strip()
        
        # Copy number of floors
        if num_floors is not None:
            ws_model["K36"] = num_floors
        
        logger.info(f"[v0] Saving modified workbook...")
        # Save the modified workbook to bytes
        output = io.BytesIO()
        model_wb.save(output)
        output.seek(0)
        
        # Close workbooks
        input_wb.close()
        model_wb.close()
        
        logger.info(f"[v0] Successfully processed input file with address: {address_clean}")
        
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"[v0] Error in process_bespoke_model: {str(e)}")
        raise

@app.get("/")
async def root():
    return {
        "message": "Bespoke Financial Model API", 
        "status": "running",
        "template_exists": os.path.exists(TEMPLATE_MODEL_PATH)
    }

@app.post("/api/process")
async def process_file(file: UploadFile = File(...)):
    """
    Process uploaded Bespoke Input Sheet and return the completed model
    """
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")
        
        # Validate file type
        valid_extensions = ('.xlsx', '.xlsm')
        if not file.filename.lower().endswith(valid_extensions):
            raise HTTPException(
                status_code=400, 
                detail="Invalid file type. Please upload a Bespoke Input Sheet (.xlsx or .xlsm)"
            )
        
        logger.info(f"[v0] Processing file: {file.filename}")
        
        # Read the uploaded file
        contents = await file.read()
        
        if len(contents) == 0:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")
        
        logger.info(f"[v0] File size: {len(contents)} bytes")
        
        # Process the file
        try:
            output_bytes = process_bespoke_model(contents)
        except FileNotFoundError as e:
            logger.error(f"[v0] Template file not found: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))
        except ValueError as e:
            logger.error(f"[v0] Validation error: {str(e)}")
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            logger.error(f"[v0] Error processing file: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
        
        # Save to temporary file
        temp_dir = tempfile.gettempdir()
        output_filename = f"Bespoke Model - US - v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
        output_path = os.path.join(temp_dir, output_filename)
        
        with open(output_path, 'wb') as f:
            f.write(output_bytes)
        
        logger.info(f"[v0] Successfully generated model: {output_filename} ({len(output_bytes)} bytes)")
        
        # Return success response
        return {
            "success": True,
            "download_filename": output_filename,
            "message": "Bespoke Model generated successfully"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[v0] Unexpected error processing file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """
    Download the generated Excel file
    """
    try:
        # Validate filename to prevent directory traversal
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found or has expired")
        
        logger.info(f"Downloading file: {filename}")
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
