import os
import re
import shutil
from openpyxl import load_workbook

def generate_model(input_path):
    base_dir = os.path.dirname(input_path)

    # --- Find model file ---
    model_file = os.path.join(base_dir, "Scripts and Models", "Bespoke Model - US - v2.xlsm")
    if not os.path.exists(model_file):
        raise FileNotFoundError("❌ Model file not found.")

    # --- Open input workbook ---
    wb_input = load_workbook(input_path, data_only=True)
    ws_input = wb_input["Sales Team Input Sheet"]

    # --- Read values from the input sheet ---
    def get_value(cell):
        val = ws_input[cell].value
        return str(val).strip() if val else ""

    address_raw = get_value("F7")
    additional_info = get_value("F9")
    latitude = ws_input["F13"].value
    longitude = ws_input["F15"].value
    sqft = ws_input["F29"].value
    market_rent = ws_input["F37"].value
    yes_no_value = get_value("F54")
    num_floors = ws_input["F56"].value

    if not address_raw:
        raise ValueError("Address in F7 is empty. Please fill it in.")

    address_clean = address_raw.replace("Dr", "Drive").replace("Blvd", "Boulevard")

    folder_name = f"{address_clean} {additional_info}"
    invalid_chars = '<>:"/\\|?*'
    for ch in invalid_chars:
        folder_name = folder_name.replace(ch, "_")

    new_folder = os.path.join(base_dir, folder_name)
    os.makedirs(new_folder, exist_ok=True)

    # --- Copy model workbook ---
    model_filename = "".join("_" if ch in invalid_chars else ch for ch in address_clean)
    new_file = os.path.join(new_folder, f"{model_filename}.xlsm")
    shutil.copy2(model_file, new_file)

    # --- Open the model workbook ---
    wb_model = load_workbook(new_file)
    ws_model = wb_model["Sales Team Input Sheet"]

    # --- Write values ---
    ws_model["E6"].value = address_clean
    ws_model["E12"].value = latitude
    ws_model["E14"].value = longitude
    ws_model["E34"].value = sqft

    # --- Handle market rent (dropdown logic) ---
    if market_rent:
        try:
            val = float(market_rent)
            if val == 15:
                ws_model["K10"].value = "15 - 20"
            elif val == 20:
                ws_model["K10"].value = "20 - 25"
            else:
                ws_model["K10"].value = str(market_rent)
        except Exception:
            ws_model["K10"].value = str(market_rent)

    ws_model["K34"].value = yes_no_value
    ws_model["K36"].value = num_floors

    # --- Save ---
    wb_model.save(new_file)
    return new_file


if __name__ == "__main__":
    import sys
    input_path = sys.argv[1]
    output_path = generate_model(input_path)
    print(f"✅ Model generated successfully: {output_path}")
